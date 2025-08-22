from __future__ import annotations
import io, re, json
from typing import Optional, Dict, List
import numpy as np
import pandas as pd

import psycopg2
from google.cloud import storage

# ========= CONFIG =========
OUT_URI = "gs://pdi-prvrstr-stg-hcb-dev/report_stg/output_kpi_02.xlsx"
SHEET1_NAME = "summary"
SHEET2_NAME = "roster-KPIs"
CONFIG_PATH = "gs://YOUR_BUCKET/path/to/db_config.json"

# Canonical -> Friendly
FRIENDLY_HEADERS: Dict[str, str] = {
    "business_owner": "Business Team",
    "group_type": "Group Type",
    "roster_id": "Roster ID",
    "roster_name": "Provider Entity",
    "parent_transaction_type": "Parent Transaction Type",
    "transaction_type": "Transaction Type",
    "case_number": "Case Number",
    "roster_date": "Roster Date",
}

SHEET1_COLS = [
    "Business Team", "Group Type", "Roster ID", "Provider Entity",
    "Parent Transaction Type", "Transaction Type", "Case Number", "Roster Date",
    "Conformance TAT", "# of rows in", "# of rows out",
    "# of unique NPI's in Input", "# of unique NPI's in Output",
]

BUSINESS_OWNER_COL = "Business Team"
SHEET2_COLS_WITH_OWNER = [
    BUSINESS_OWNER_COL,
    "# New Roster Formats",
    "# Changed Roster Formats",
    "# of Rosters with no Set up or Format Change",
    "# Complex Rosters",
    "All Rosters",
]

# ========= I/O =========
def _load_json(path: str) -> dict:
    """Load JSON from GCS (gs://...) or local file path."""
    if path.startswith("gs://"):
        bucket, blob_path = path[5:].split("/", 1)
        text = storage.Client().bucket(bucket).blob(blob_path).download_as_text()
        return json.loads(text)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def query_df_from_config(config_path: str) -> pd.DataFrame:
    """Connect using credentials from JSON and return SQL result as DataFrame."""
    cfg = _load_json(config_path)
    sql = cfg["sql_query"]
    db = cfg["database"]
    conn = psycopg2.connect(
        dbname=db["dbname"],
        user=db["user"],
        password=db["password"],
        host=db["host"],
        port=int(db["port"]),
    )
    try:
        return pd.read_sql_query(sql, conn)
    finally:
        conn.close()

# ======== HELPERS ========
def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())

def find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """Return a column matching any candidate (normalized exact, then substring)."""
    by_norm = {_norm(c): c for c in df.columns}
    for c in candidates:
        k = _norm(c); 
        if k in by_norm: 
            return by_norm[k]
    for c in candidates:
        ck = _norm(c)
        for col in df.columns:
            if ck and ck in _norm(col):
                return col
    return None

def _business_owner_col(df: pd.DataFrame) -> Optional[str]:
    return find_col(df, ["business_owner", "Business Owner", "business owner"])

def _version_status_series(df: pd.DataFrame) -> Optional[pd.Series]:
    col = find_col(df, ["version_status", "Version Status", "status"])
    if not col:
        return None
    # SAFE: don’t force pandas StringDtype; work on Python strings
    return (
        df[col]
        .astype(object)
        .map(lambda x: "" if pd.isna(x) else str(x))
        .str.strip()
        .str.replace(r"(?i)[^\w\s\-]+", "", regex=True)
        .str.upper()
    )

def _to_text_series(s: pd.Series) -> pd.Series:
    """Robust text view for sizing/export; avoids pandas StringDtype pitfalls."""
    return s.astype(object).map(lambda v: "" if pd.isna(v) else str(v))

def add_case_and_roster_date(summary_df: pd.DataFrame, src_df: pd.DataFrame) -> pd.DataFrame:
    out = summary_df.copy().reset_index(drop=True)
    src = src_df.reset_index(drop=True)
    out["case_number"] = src["case_number"] if "case_number" in src.columns else pd.NA
    out["roster_date"] = src["roster_date"] if "roster_date" in src.columns else pd.NA
    return out

def make_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    """Remove tz from datetimes and replace NaT before writing to Excel."""
    out = df.copy()
    # Drop timezone from tz-aware datetime columns
    for col in out.select_dtypes(include=["datetimetz"]).columns:
        out[col] = out[col].dt.tz_localize(None)
    # If any object columns contain tz-aware datetimes, fix them
    for col in out.columns[out.dtypes.eq("object")]:
        s = out[col]
        try:
            has_tz = s.map(lambda x: getattr(x, "tzinfo", None) is not None).any()
        except Exception:
            has_tz = False
        if has_tz:
            out[col] = pd.to_datetime(out[col], errors="coerce", utc=True).dt.tz_localize(None)
    # Excel-friendly: replace NaT with empty string
    out = out.replace({pd.NaT: ""})
    return out

def autosize_and_freeze_openpyxl(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    """Auto-size columns and freeze header row, safely."""
    ws = writer.sheets[sheet_name]
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(df.columns, start=1):
        series = _to_text_series(df[col])
        max_len = max(series.map(len).max() if len(series) else 0, len(str(col)))
        ws.column_dimensions[get_column_letter(idx)].width = min(int(max_len) + 2, 60)
    ws.freeze_panes = "A2"

def add_new_roster_formats(df: pd.DataFrame, include=("NEW_FILE", "NEW_VERSION", "ACTIVE")) -> pd.DataFrame:
    out, owner = df.copy(), _business_owner_col(df)
    if not owner:
        out["# New Roster Formats"] = 0
        return out
    vs = _version_status_series(out)
    if vs is None:
        out["# New Roster Formats"] = 0
        return out
    statuses = [s.replace("_", " ") for s in include]
    is_new = vs.isin(statuses)
    counts = is_new.groupby(out[owner]).transform("sum")
    out["# New Roster Formats"] = counts.astype(int)
    return out

def add_changed_roster_formats(df: pd.DataFrame) -> pd.DataFrame:
    out, owner = df.copy(), _business_owner_col(df)
    if not owner:
        out["# Changed Roster Formats"] = 0
        return out
    vs = _version_status_series(out)
    if vs is None:
        out["# Changed Roster Formats"] = 0
        return out
    prev = vs.groupby(out[owner]).shift(1)
    changed = vs.eq("NEW_VERSION") & prev.notna() & prev.ne("NEW_VERSION")
    out["# Changed Roster Formats"] = changed.groupby(out[owner]).transform("sum").astype(int)
    return out

def add_no_setup_or_format_change(df: pd.DataFrame,
    colname: str = "# of Rosters with no Set up or Format Change") -> pd.DataFrame:
    out, owner = df.copy(), _business_owner_col(df)
    if not owner:
        out[colname] = 0
        return out
    uniq = out[owner].groupby(out[owner]).transform(lambda s: s.nunique(dropna=True))
    out[colname] = (uniq < 1).astype(int)
    return out

def add_complex_rosters(df: pd.DataFrame) -> pd.DataFrame:
    out, owner = df.copy(), _business_owner_col(df)
    if not owner:
        out["# Complex Rosters"] = 0
        return out
    cx_col = find_col(out, ["complexity", "Complexity", "complex"])
    if not cx_col:
        out["# Complex Rosters"] = 0
        return out
    is_cx = _to_text_series(out[cx_col]).str.upper().str.contains("COMPLEX", na=False)
    out["# Complex Rosters"] = is_cx.groupby(out[owner]).transform("sum").astype(int)
    return out

def add_all_rosters(df: pd.DataFrame) -> pd.DataFrame:
    out, owner = df.copy(), _business_owner_col(df)
    if not owner:
        out["All Rosters"] = 1
        return out
    vc = out[owner].value_counts(dropna=False)
    out["All Rosters"] = out[owner].map(vc).astype(int)
    return out

def add_conformance_tat(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    start_candidates = ["prms_posted_timestamp"]
    end_candidates = ["file_ingestion_timestamp"]
    lower_map = {c.lower(): c for c in out.columns}
    start_cols = [lower_map[n.lower()] for n in start_candidates if n.lower() in lower_map]
    end_cols = [lower_map[n.lower()] for n in end_candidates if n.lower() in lower_map]
    if not start_cols or not end_cols:
        out["Conformance TAT"] = None
        return out
    for c in start_cols + end_cols:
        out[c] = pd.to_datetime(out[c], errors="coerce", utc=True)
    start_ts = out[start_cols].bfill(axis=1).iloc[:, 0]
    end_ts = out[end_cols].bfill(axis=1).iloc[:, 0]
    valid = start_ts.notna() & end_ts.notna()
    tat = pd.Series(pd.NaT, index=out.index, dtype="timedelta64[ns]")
    tat.loc[valid] = (end_ts[valid] - start_ts[valid]).dt.round("S")
    tat = tat.mask(tat < pd.Timedelta(0), pd.Timedelta(0))
    comps = tat.dt.components
    tat_str = (
        comps["days"].astype("Int64").astype(str).str.zfill(2) + "/" +
        comps["hours"].astype("Int64").astype(str).str.zfill(2) + "/" +
        comps["minutes"].astype("Int64").astype(str).str.zfill(2) + "/" +
        comps["seconds"].astype("Int64").astype(str).str.zfill(2)
    ).where(tat.notna(), None)
    out["Conformance TAT"] = tat_str
    return out

def add_rows_counts(df_src: pd.DataFrame, df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    col_in  = find_col(df_src, ["input_rec_count", "input records count"])
    col_out = find_col(df_src, ["conformed_rec_count", "output_rec_count", "output records count"])
    out["# of rows in"]  = df_src[col_in]  if col_in  is not None and len(df_src) == len(out) else len(df_src)
    out["# of rows out"] = df_src[col_out] if col_out is not None and len(df_src) == len(out) else len(out)
    return out

def add_unique_npi_counts(df_src: pd.DataFrame, df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    in_cnt  = find_col(df_src, ["input_unique_npi_count", "unique npi input", "unique npi count input"])
    out_cnt = find_col(df_src, ["conformed_unique_npi_count", "unique npi output", "unique npi count output"])
    if in_cnt and len(df_src) == len(out):
        out["# of unique NPI's in Input"] = df_src[in_cnt]
    else:
        npi_in = find_col(df_src, ["npi", "npi_number", "npi id", "npiid"])
        out["# of unique NPI's in Input"] = df_src[npi_in].nunique(dropna=True) if npi_in else pd.NA
    if out_cnt and len(df_src) == len(out):
        out["# of unique NPI's in Output"] = df_src[out_cnt]
    else:
        npi_out = find_col(df_src, ["npi", "npi_number", "npi id", "npiid"])
        out["# of unique NPI's in Output"] = df_src[npi_out].nunique(dropna=True) if npi_out else pd.NA
    return out

def apply_friendly_headers(df: pd.DataFrame, columns_mapping: Dict[str, str]) -> pd.DataFrame:
    """Rename mapped columns and KEEP KPI columns. Ensures 'All Rosters' is retained."""
    if not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    df2 = df.rename(columns=columns_mapping)
    friendly_cols = [columns_mapping[c] for c in columns_mapping if c in df.columns]
    kpi_cols = [c for c in df2.columns if c.startswith("#") or c in ("Conformance TAT", "All Rosters")]
    keep_cols = [c for c in (friendly_cols + kpi_cols) if c in df2.columns]
    return df2[keep_cols]

# ----------------------------- Excel utilities -----------------------------
def autosize_and_freeze(ws, df: pd.DataFrame):
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(df.columns, start=1):
        series = _to_text_series(df[col])
        max_len = max(series.map(len).max() if len(series) else 0, len(str(col)))
        ws.column_dimensions[get_column_letter(idx)].width = min(int(max_len) + 2, 60)
    ws.freeze_panes = "A2"

def write_two_sheet_excel_gcs(df: pd.DataFrame, gcs_uri: str) -> None:
    """Sheet1 detail, Sheet2 aggregated by Business Team, upload to GCS."""
    # Sheet 1
    df1 = df.reindex(columns=[c for c in SHEET1_COLS if c in df.columns])

    # Sheet 2
    missing_owner = BUSINESS_OWNER_COL not in df.columns
    kpis_present = [c for c in SHEET2_COLS_WITH_OWNER if c in df.columns and c != BUSINESS_OWNER_COL]
    if missing_owner:
        df2 = pd.DataFrame(columns=SHEET2_COLS_WITH_OWNER)
    else:
        if kpis_present:
            df2 = (
                df[[BUSINESS_OWNER_COL] + kpis_present]
                .drop_duplicates(subset=[BUSINESS_OWNER_COL])
                .reindex(columns=[c for c in SHEET2_COLS_WITH_OWNER if c in ([BUSINESS_OWNER_COL] + kpis_present)])
            )
        else:
            df2 = df[[BUSINESS_OWNER_COL]].drop_duplicates()
            for k in SHEET2_COLS_WITH_OWNER:
                if k != BUSINESS_OWNER_COL and k not in df2.columns:
                    df2[k] = pd.NA
            df2 = df2[SHEET2_COLS_WITH_OWNER]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if not df1.empty:
            df1.to_excel(writer, sheet_name=SHEET1_NAME, index=False)
            autosize_and_freeze_openpyxl(writer, df1, SHEET1_NAME)
        else:
            pd.DataFrame(columns=SHEET1_COLS).to_excel(writer, sheet_name=SHEET1_NAME, index=False)
            autosize_and_freeze_openpyxl(writer, pd.DataFrame(columns=SHEET1_COLS), SHEET1_NAME)

        df2.to_excel(writer, sheet_name=SHEET2_NAME, index=False)
        autosize_and_freeze_openpyxl(writer, df2, SHEET2_NAME)

    buf.seek(0)

    if not gcs_uri.startswith("gs://"):
        raise ValueError(f"gcs_uri must start with gs://, got: {gcs_uri}")
    bucket_name, blob_path = gcs_uri[5:].split("/", 1)
    storage.Client().bucket(bucket_name).blob(blob_path).upload_from_file(
        buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ===================== MAIN PIPELINE =====================
def run_pipeline_to_gcs(
    src_df: Optional[pd.DataFrame] = None,
    out_uri: str = OUT_URI,
    config_path: Optional[str] = None,
) -> pd.DataFrame:
    """
    Runs the full KPI pipeline and uploads the Excel to GCS.
    - If src_df is None, it will read using CONFIG JSON (config_path or CONFIG_PATH).
    """
    if src_df is None:
        cfg = config_path or CONFIG_PATH
        src = query_df_from_config(cfg)
    else:
        src = src_df.copy()

    # === KPI build sequence ===
    df = src.copy()
    df = add_new_roster_formats(df)
    df = add_changed_roster_formats(df)
    df = add_no_setup_or_format_change(df)
    df = add_case_and_roster_date(df, src)
    df = add_complex_rosters(df)
    df = add_all_rosters(df)
    df = add_conformance_tat(df)
    df = add_rows_counts(src, df)
    df = add_unique_npi_counts(src, df)

    # Friendly headers then excel-safe
    df = apply_friendly_headers(df, FRIENDLY_HEADERS)
    df = make_excel_safe(df)

    # Write two-sheet Excel and upload
    write_two_sheet_excel_gcs(df, out_uri)
    return df

**************************************************************

# file: simple_db_query_min.py
import json
import pandas as pd
import psycopg2
from google.cloud import storage


def _load_json(path: str) -> dict:
    """Load JSON from GCS (gs://...) or local file path."""
    if path.startswith("gs://"):
        bucket, blob_path = path[5:].split("/", 1)
        client = storage.Client()
        text = client.bucket(bucket).blob(blob_path).download_as_text()
        return json.loads(text)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def query_df_from_config(config_path: str) -> pd.DataFrame:
    """Connect using credentials from JSON and return the SQL result as a DataFrame."""
    cfg = _load_json(config_path)

    # Required fields only
    sql = cfg["sql_query"]
    db = cfg["database"]
    conn = psycopg2.connect(
        dbname=db["dbname"],
        user=db["user"],
        password=db["password"],
        host=db["host"],
        port=int(db["port"]),
    )
    try:
        return pd.read_sql_query(sql, conn)
    finally:
        conn.close()


if __name__ == "__main__":
    CONFIG_PATH = "gs://YOUR_BUCKET/path/to/db_config.json"  # or local path
    df = query_df_from_config(CONFIG_PATH)
    print(df.head())
    print("Rows:", len(df))

**************************************
# file: simple_db_query_from_config.py
import json
import pandas as pd
import psycopg2

# If your CONFIG_PATH is a gs:// URI, this import is used
from google.cloud import storage


CONFIG_PATH = "gs://YOUR_BUCKET/path/to/db_config.json"  # or "/path/local/db_config.json"


def load_config(path: str) -> dict:
    """Load JSON config from GCS (gs://...) or local filesystem."""
    if path.startswith("gs://"):
        bucket_name, blob_path = path[5:].split("/", 1)
        client = storage.Client()
        text = client.bucket(bucket_name).blob(blob_path).download_as_text()
        return json.loads(text)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def query_df_from_config(config_path: str) -> pd.DataFrame:
    """
    Expects JSON like:
    {
      "sql_query": "SELECT ...",
      "database": {
        "dbname": "...",
        "user": "...",
        "password": "...",
        "host": "...",
        "port": 5432
      }
    }
    (If "database" is omitted, top-level keys are used.)
    """
    cfg = load_config(config_path)

    # Pull SQL
    sql = cfg.get("sql_query") or cfg.get("SQL") or "SELECT 1"

    # Pull DB creds (nested under "database" or top-level)
    db_cfg = cfg.get("database", cfg)
    dbname = db_cfg["dbname"]
    user = db_cfg["user"]
    password = db_cfg["password"]
    host = db_cfg["host"]
    port = int(db_cfg.get("port", 5432))

    # Connect & query
    with psycopg2.connect(
        dbname=dbname,
        user=user,
        password=password,
        host=host,
        port=port,
        # If needed: sslmode="require"
    ) as conn:
        df = pd.read_sql_query(sql, conn)

    return df


if __name__ == "__main__":
    df = query_df_from_config(CONFIG_PATH)
    print(df.head())
    print("Rows:", len(df))
